import os
import datetime
import threading
from openpyxl import load_workbook, Workbook

DATA_DIR = os.environ.get("DATA_DIR", ".")
EXCEL_FILE = os.path.join(DATA_DIR, "bookings.xlsx")
lock = threading.Lock()

def get_location_code(loc_val):
    """Normalizes location values (Ukrainian display names or codes) into standard keys (kyiv, chabany, obukhiv)."""
    if not loc_val:
        return ""
    loc_str = str(loc_val).strip().lower()
    if "оболонь" in loc_str or "київ" in loc_str or "kyiv" in loc_str:
        return "kyiv"
    elif "чабани" in loc_str or "chabany" in loc_str:
        return "chabany"
    elif "обухів" in loc_str or "obukhiv" in loc_str:
        return "obukhiv"
    return loc_str

def init_excel():
    """Initializes the Excel workbook with headers if it does not exist, or updates headers if necessary."""
    # Check if the workbook is in old format and rename it
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            # If the cell A1 is "Booking ID" (old format), we backup and delete
            if ws.cell(row=1, column=1).value == "Booking ID":
                wb.close()
                backup_name = os.path.join(DATA_DIR, "bookings_backup.xlsx")
                # Find an unused backup name
                counter = 1
                while os.path.exists(backup_name):
                    backup_name = os.path.join(DATA_DIR, f"bookings_backup_{counter}.xlsx")
                    counter += 1
                os.rename(EXCEL_FILE, backup_name)
                print(f"Backed up old Excel file to {backup_name}")
            else:
                wb.close()
        except Exception as e:
            print(f"Error checking Excel format: {e}")

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Row 1 values as requested in user's layout screenshot
        row1 = ["Час", "Телеграм", "@ggme", "380689630697", "Обухів", 9, 9, 9, 9]
        ws.append(row1)
        
        # Row 2 headers matching the Ukrainian labels + tracking headers
        row2 = [
            "локація",                      # Col 1 (A)
            "ім'я клієнта",                 # Col 2 (B)
            "прізвище клієнта",             # Col 3 (C)
            "телефон",                      # Col 4 (D)
            "телеграм юзернейм",            # Col 5 (E)
            "телеграм ID",                  # Col 6 (F)
            "дата та час",                  # Col 7 (G) - formatted like: 21.05.2026 14:00
            "ім'я іменинника",              # Col 8 (H)
            "вік іменинника",               # Col 9 (I)
            "кількість дітей",              # Col 10 (J)
            "вік дітей (range)",            # Col 11 (K)
            "які послуги цікавлять",        # Col 12 (L)
            "коментарі/особливі побажання", # Col 13 (M)
            "Booking ID",                   # Col 14 (N)
            "Total Price",                  # Col 15 (O)
            "Status",                       # Col 16 (P)
            "Timestamp"                     # Col 17 (Q)
        ]
        ws.append(row2)
        wb.save(EXCEL_FILE)
        print(f"Initialized new Excel file matching user layout: {EXCEL_FILE}")

def book_items(location, name, lastname, phone, telegram, telegram_id, date_str, time_slot_str, items_list, total_price,
               birthday_name="", birthday_age="", kids_count="", kids_age_range="", comments="", status="Очікує дзвінка"):
    """
    Attempts to book a list of items for a specific location, date, and time slot.
    Performs conflict checks using the new column indices. If no conflicts, writes to bookings.xlsx.
    
    Returns:
        (success (bool), message (str), booking_id (str))
    """
    with lock:
        init_excel()
        
        # Open workbook
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception as e:
            return False, f"Помилка при зчитуванні бази даних Excel: {str(e)}", None

        # Clean/Normalize inputs
        loc_code = get_location_code(location)
        
        # Convert YYYY-MM-DD to DD.MM.YYYY
        try:
            parts = date_str.split("-")
            if len(parts) == 3:
                date_formatted = f"{parts[2]}.{parts[1]}.{parts[0]}"
            else:
                date_formatted = date_str
        except Exception:
            date_formatted = date_str
            
        time_normalized = str(time_slot_str).strip()
        datetime_formatted = f"{date_formatted} {time_normalized}"
        
        # Build set of items we want to book
        requested_items = {str(item).strip().lower() for item in items_list}
        
        # 1. Conflict Check
        conflicts = []
        
        # Iterate starting from row 3 (skip header row 1 and row 2)
        for row in range(3, ws.max_row + 1):
            row_booking_id = ws.cell(row=row, column=14).value
            row_location = ws.cell(row=row, column=1).value
            row_datetime = ws.cell(row=row, column=7).value
            row_items_raw = ws.cell(row=row, column=12).value
            
            # Skip empty or corrupted rows
            if not row_booking_id or not row_location or not row_datetime or not row_items_raw:
                continue
                
            # Normalize cell values for comparison using our standardized helper
            row_loc_code = get_location_code(row_location)
            row_datetime_norm = str(row_datetime).strip()
            
            # Check if this booking conflicts (same location, same date-time)
            if row_loc_code == loc_code and row_datetime_norm == datetime_formatted:
                # Split items list by comma
                booked_items = {item.strip().lower() for item in str(row_items_raw).split(",")}
                
                # Check for overlap
                overlap = requested_items.intersection(booked_items)
                if overlap:
                    # Find matching exact names for error reporting
                    for item in items_list:
                        if item.strip().lower() in overlap:
                            conflicts.append(item)

        if conflicts:
            wb.close()
            return False, f"Ці позиції вже заброньовані на цей час: {', '.join(conflicts)}", None

        # 2. Determine Booking ID from Column 14 (N)
        last_id_num = 1000
        for row in range(3, ws.max_row + 1):
            val = ws.cell(row=row, column=14).value
            if val and str(val).startswith("EP-"):
                try:
                    num = int(str(val).split("-")[1])
                    if num > last_id_num:
                        last_id_num = num
                except ValueError:
                    pass
        new_booking_id = f"EP-{last_id_num + 1}"

        # 3. Add Booking
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        items_joined = ", ".join(items_list)
        
        # Translate location code to beautiful name
        loc_display = "Оболонь" if loc_code == "kyiv" else ("Чабани" if loc_code == "chabany" else "Обухів")
        
        new_row = [
            loc_display,                    # Col 1 (A)
            name,                           # Col 2 (B)
            lastname,                       # Col 3 (C)
            phone,                          # Col 4 (D)
            telegram if telegram.startswith("@") or not telegram else f"@{telegram}", # Col 5 (E)
            telegram_id,                    # Col 6 (F)
            datetime_formatted,             # Col 7 (G)
            birthday_name,                  # Col 8 (H)
            birthday_age,                   # Col 9 (I)
            kids_count,                     # Col 10 (J)
            kids_age_range,                 # Col 11 (K)
            items_joined,                   # Col 12 (L)
            comments,                       # Col 13 (M)
            new_booking_id,                 # Col 14 (N)
            f"{total_price} грн",           # Col 15 (O)
            status,                         # Col 16 (P)
            timestamp                       # Col 17 (Q)
        ]
        
        try:
            ws.append(new_row)
            wb.save(EXCEL_FILE)
            wb.close()
            return True, "Бронювання успішно створено!", new_booking_id
        except Exception as e:
            return False, f"Не вдалося зберегти запис у Excel: {str(e)}", None

def get_booked_slots(location, date_str):
    """
    Returns a dictionary of blocked time slots and the items booked in them.
    Used by the frontend to see what is already busy.
    Returns: { "12:00": ["Item A", "Item B"], "13:00": [...] }
    """
    with lock:
        init_excel()
        slots_map = {}
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception:
            return slots_map
            
        loc_code = get_location_code(location)
        
        # Convert YYYY-MM-DD to DD.MM.YYYY
        try:
            parts = date_str.split("-")
            if len(parts) == 3:
                date_formatted = f"{parts[2]}.{parts[1]}.{parts[0]}"
            else:
                date_formatted = date_str
        except Exception:
            date_formatted = date_str
        
        for row in range(3, ws.max_row + 1):
            row_loc_raw = ws.cell(row=row, column=1).value
            row_datetime = ws.cell(row=row, column=7).value
            row_items = ws.cell(row=row, column=12).value
            
            if not row_loc_raw or not row_datetime or not row_items:
                continue
                
            row_loc_code = get_location_code(row_loc_raw)
                
            if row_loc_code == loc_code:
                datetime_str = str(row_datetime).strip()
                # datetime_str is like "22.05.2026 14:00"
                if " " in datetime_str:
                    date_part, time_part = datetime_str.split(" ", 1)
                    if date_part.strip() == date_formatted:
                        time_str = time_part.strip()
                        item_list = [item.strip() for item in str(row_items).split(",")]
                        if time_str not in slots_map:
                            slots_map[time_str] = []
                        slots_map[time_str].extend(item_list)
                
        wb.close()
        return slots_map

def get_all_bookings():
    """
    Returns all bookings stored in bookings.xlsx as a list of dicts.
    Useful for the admin panel.
    """
    with lock:
        init_excel()
        bookings = []
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception:
            return bookings
            
        for row in range(3, ws.max_row + 1):
            location_raw = ws.cell(row=row, column=1).value
            customer_name = ws.cell(row=row, column=2).value
            last_name = ws.cell(row=row, column=3).value
            phone = ws.cell(row=row, column=4).value
            telegram = ws.cell(row=row, column=5).value
            telegram_id = ws.cell(row=row, column=6).value
            datetime_raw = ws.cell(row=row, column=7).value
            birthday_name = ws.cell(row=row, column=8).value
            birthday_age = ws.cell(row=row, column=9).value
            kids_count = ws.cell(row=row, column=10).value
            kids_age_range = ws.cell(row=row, column=11).value
            items = ws.cell(row=row, column=12).value
            comments = ws.cell(row=row, column=13).value
            booking_id = ws.cell(row=row, column=14).value
            total_price = ws.cell(row=row, column=15).value
            status = ws.cell(row=row, column=16).value
            timestamp = ws.cell(row=row, column=17).value
            
            if not booking_id:
                continue
                
            if not status:
                status = "Очікує дзвінка"
                
            # Split G (дата та час) into date and time
            date_val = ""
            time_slot_val = ""
            if datetime_raw:
                parts = str(datetime_raw).split(" ", 1)
                if len(parts) == 2:
                    # Convert DD.MM.YYYY back to YYYY-MM-DD for UI components
                    dd_mm_yyyy = parts[0].split(".")
                    if len(dd_mm_yyyy) == 3:
                        date_val = f"{dd_mm_yyyy[2]}-{dd_mm_yyyy[1]}-{dd_mm_yyyy[0]}"
                    else:
                        date_val = parts[0]
                    time_slot_val = parts[1]
                else:
                    date_val = str(datetime_raw)
            
            loc_code = get_location_code(location_raw)
                
            bookings.append({
                "booking_id": str(booking_id),
                "timestamp": str(timestamp) if timestamp else "",
                "location": loc_code,
                "location_display": str(location_raw) if location_raw else "",
                "customer_name": f"{customer_name} {last_name}".strip() if last_name else (str(customer_name) if customer_name else ""),
                "first_name": str(customer_name) if customer_name else "",
                "last_name": str(last_name) if last_name else "",
                "phone": str(phone) if phone else "",
                "telegram": str(telegram) if telegram else "",
                "telegram_id": str(telegram_id) if telegram_id else "",
                "date": date_val,
                "time_slot": time_slot_val,
                "items": [i.strip() for i in str(items).split(",")] if items else [],
                "total_price": str(total_price) if total_price else "",
                "status": str(status),
                "birthday_name": str(birthday_name) if birthday_name else "",
                "birthday_age": str(birthday_age) if birthday_age else "",
                "kids_count": str(kids_count) if kids_count else "",
                "kids_age_range": str(kids_age_range) if kids_age_range else "",
                "comments": str(comments) if comments else ""
            })
            
        wb.close()
        # Sort bookings by date and time slot descending
        try:
            bookings.sort(key=lambda x: (x["date"], x["time_slot"]), reverse=True)
        except Exception:
            pass
        return bookings

def update_booking_status(booking_id, new_status):
    """
    Updates the status of a specific booking in bookings.xlsx (Column 16).
    """
    with lock:
        init_excel()
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception as e:
            return False, f"Помилка відкриття файлу: {str(e)}"
            
        booking_id_norm = str(booking_id).strip().upper()
        found = False
        
        for row in range(3, ws.max_row + 1):
            val = ws.cell(row=row, column=14).value
            if val and str(val).strip().upper() == booking_id_norm:
                ws.cell(row=row, column=16, value=str(new_status))
                found = True
                break
                
        if not found:
            wb.close()
            return False, "Бронювання не знайдено"
            
        try:
            wb.save(EXCEL_FILE)
            wb.close()
            return True, "Статус успішно оновлено!"
        except Exception as e:
            return False, f"Не вдалося зберегти оновлення: {str(e)}"
